"""
muavin_pipeline.py — n8n Entegrasyon Giriş Noktası
====================================================
Kullanım:
  python muavin_pipeline.py --input "dosya.xlsx" --output "cikti.xlsx" --sablon "sablon.xlsx"

n8n Execute Command node örneği:
  python "d:/FINANS_AI/01_RAW/muavin_pipeline.py"
    --input  "{{ $json.input_path }}"
    --output "{{ $json.output_path }}"
    --sablon "d:/FINANS_AI/01_RAW/Örnek Sablonlar.xlsx"

Çıktı: stdout'a JSON yazar (n8n'in okuyacağı veri)
{
  "status": "success",
  "format": "VNK | WOW",
  "input_path": "...",
  "output_path": "...",
  "rows_in": 478,
  "rows_out": 162,
  "leaf_accounts": ["191.01.003", ...],
  "toplam_borc": 764980.15,
  "toplam_alacak": 764980.15,
  "log_path": "...",
  "error": null
}
"""

import sys
import re
import json
import shutil
import logging
import argparse
import traceback
from io import StringIO
from pathlib import Path
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta

try:
    import xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False

# ─────────────────────────────────────────────
# SABITLER
# ─────────────────────────────────────────────
KAYNAK_SISTEM = "Logo"
DOVIZ_KURU    = 1.0

# ─────────────────────────────────────────────
# LOGGING — hem dosyaya hem belleğe yaz
# ─────────────────────────────────────────────
_log_buffer = StringIO()

def _setup_logging(log_path: Path) -> logging.Logger:
    fmt = "%(asctime)s | %(levelname)s | %(message)s"
    logger = logging.getLogger("muavin_pipeline")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.addHandler(logging.FileHandler(log_path, encoding="utf-8"))
    sh = logging.StreamHandler(_log_buffer)
    sh.setFormatter(logging.Formatter(fmt))
    logger.addHandler(sh)
    return logger


# ─────────────────────────────────────────────
# FORMAT TESPİTİ
# ─────────────────────────────────────────────
def format_tespit(path: Path) -> str:
    """
    Dosyanın muavin formatını tespit eder.

    Dönüş:
      "VNK"  — Tek hesaplı, düz liste: R1="Muavin Defter [...]", R2=kolon başlıkları
      "WOW"  — Çok hesaplı, blok yapı: R1="MUAVİN DEFTER", hesap kodları satır başında
      "UNKNOWN"
    """
    rows, _ = _lese_rows(path)
    rows = rows[:6]

    r1 = str(rows[0][0] or "").upper()

    # VNK tipi: başlık satırı genellikle MUAVİN SON DÖKÜM içerir
    if "SON D" in r1 or "DÖKÜM" in r1 or "DOKUМ" in r1:
        return "VNK"

    # R2 kolonlarına bak
    if len(rows) > 1:
        r2_cols = [str(c or "").upper() for c in rows[1]]
        if "HESAP KODU" in r2_cols:
            return "VNK"

    # WOW tipi: birkaç satır içinde "hesap_kodu ad" formatı var
    _re = re.compile(r"^\d{3}(?:\.\d{2,3})+\s+\S")
    for row in rows[2:]:
        if row[0] and _re.match(str(row[0]).strip()):
            return "WOW"

    # Son şans: R1 "MUAVİN DEFTER" ile başlıyor
    if r1.startswith("MUAVİN") or r1.startswith("MUAVIN"):
        return "WOW"

    return "UNKNOWN"


def _lese_rows(path: Path) -> tuple:
    """
    Herhangi bir Excel dosyasını okur, satır listesi döner.
    Desteklenen formatlar:
      - .xlsx (openpyxl)
      - .xls görünümlü .xlsx (openpyxl, kopyalanarak)
      - .xls gerçek BIFF (xlrd) — tarihler datetime'a çevrilir
    Dönüş: (rows: list[tuple], sheet_name: str)
    """
    # 1) openpyxl direkt
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True)), ws.title
    except Exception:
        pass

    # 2) xlsx-görünümlü-xls: uzantıyı değiştirerek dene
    tmp = path.with_suffix(".tmp_pipeline.xlsx")
    try:
        shutil.copy(path, tmp)
        wb = load_workbook(tmp, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        tmp.unlink(missing_ok=True)
        return rows, ws.title
    except Exception:
        tmp.unlink(missing_ok=True)

    # 3) Gerçek BIFF .xls → xlrd
    if not _XLRD_OK:
        raise ValueError(f"Gerçek .xls dosyası okunmak isteniyor ama xlrd kurulu değil: {path.name}")

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    _BASE = datetime(1899, 12, 30)
    rows = []
    for i in range(ws.nrows):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(_BASE + timedelta(days=cell.value))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(tuple(row))
    return rows, ws.name


# ─────────────────────────────────────────────
# ORTAK YARDIMCI FONKSİYONLAR
# ─────────────────────────────────────────────
def leaf_hesap_bul(kodlar: set) -> set:
    leaf = set()
    for kod in kodlar:
        prefix = str(kod).strip() + "."
        if not any(str(k).startswith(prefix) for k in kodlar if k != kod):
            leaf.add(str(kod).strip())
    return leaf


def hesap_tipi(tam_kod: str) -> str:
    m = {"1": "Aktif", "2": "Aktif", "3": "Pasif", "4": "Pasif",
         "5": "Özkaynak", "6": "Gelir", "7": "Maliyet", "8": "Maliyet", "9": "Nazım"}
    s = str(tam_kod).strip()[0] if tam_kod else ""
    return m.get(s, "")


_EXCEL_BASE = datetime(1899, 12, 30)

def tarih_temizle(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    # Excel serial number (xlrd'den gelen float tarih, örn: 46027.0)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = float(v)
        if 1000 < n < 100000:          # makul Excel serial aralığı
            return _EXCEL_BASE + timedelta(days=n)
        return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, datetime) else datetime.combine(v, datetime.min.time())
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.split(" ")[0], fmt.split(" ")[0])
        except Exception:
            pass
    try:
        return pd.to_datetime(s, dayfirst=True, errors="coerce")
    except Exception:
        return None


def sayi_temizle(v) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except Exception:
        return 0.0


# ─────────────────────────────────────────────
# VNK FORMAT — HAM VERİ OKUMA
# ─────────────────────────────────────────────
_HESAP_RE  = re.compile(r"^\d{3}(?:\.\d{2,3})*(?:\.[A-Z0-9]+)*\s*$", re.IGNORECASE)
_TARIH_RE  = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
_SKIP_VNK  = ("TARİH", "HESAP KODU")


def _is_hesap_kodu(s: str) -> bool:
    return bool(_HESAP_RE.match(s.strip()))


def _is_tarih_str(s: str) -> bool:
    return bool(_TARIH_RE.match(s.strip()))


def _parse_aciklama_vnk(raw: str) -> dict:
    r = {"belge_turu": "", "aciklama_temiz": "", "karsi_hesap_kodu": ""}
    if not raw or pd.isna(raw):
        return r
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    if not parts:
        return r
    karsi = next((p.strip() for p in reversed(parts) if _is_hesap_kodu(p)), "")
    r["karsi_hesap_kodu"] = karsi
    bilinen = ["Alınan Hizmet Faturası", "Alınan Mal Faturası", "Satış Faturası",
               "Alacak Dekontu", "Borç Dekontu", "Gider Fişi", "Mahsup Fişi",
               "Tahsilat Fişi", "Ödeme Fişi", "Storno", "Açılış Kaydı"]
    r["belge_turu"] = next(
        (b for b in bilinen for p in parts if b.lower() in p.lower()), ""
    )
    if not r["belge_turu"]:
        r["belge_turu"] = next(
            (p for p in parts if any(k in p for k in ["Fatura", "Dekont", "Fiş"])), ""
        )
    ignore = {1} if len(parts) >= 2 and parts[0] == parts[1] else set()
    clean = [p for i, p in enumerate(parts)
             if i not in ignore
             and not _is_hesap_kodu(p)
             and not _is_tarih_str(p)
             and not re.match(r"^(BT:|ISBT:|UPA:|faturaturu:|YAZI_ILE:)", p, re.I)]
    r["aciklama_temiz"] = " | ".join(clean[:6])
    return r


def vnk_oku(path: Path, log) -> pd.DataFrame:
    log.info(f"[VNK] Okuma: {path.name}")
    rows, _ = _lese_rows(path)
    header  = rows[1]
    df = pd.DataFrame(rows[2:], columns=header)

    # Kolon normalize — isim bazlı + konum bazlı fallback
    fallback = {0: "Hesap Kodu", 1: "Hesap Adı", 2: "Hesap Özel Kodu",
                3: "Hesap Yetki Kodu", 4: "Birim", 5: "Tarih",
                6: "Fiş Türü", 7: "Fiş No", 8: "Açıklama",
                9: "Borç", 10: "Alacak", 11: "Bakiye",
                12: "Doviz Kuru", 13: "FCY Tutar", 14: "FCY Bakiye",
                15: "Doviz", 16: "ID Kuru", 17: "ID Tutar", 18: "ID Bakiye"}
    rename = {}
    for i, c in enumerate(df.columns):
        s = str(c or "")
        if   "Hesap Kodu" in s:                       rename[c] = "Hesap Kodu"
        elif "Hesap Ad"   in s:                       rename[c] = "Hesap Adı"
        elif s.strip() == "Tarih":                    rename[c] = "Tarih"
        elif "Fi" in s and "T" in s and len(s) < 12: rename[c] = "Fiş Türü"
        elif "No" in s and "Fi" in s:                rename[c] = "Fiş No"
        elif "klama" in s.lower():                   rename[c] = "Açıklama"
        elif s.strip() == "Borç":                    rename[c] = "Borç"
        elif s.strip() == "Alacak":                  rename[c] = "Alacak"
        elif s.strip() == "Bakiye":                  rename[c] = "Bakiye"
        elif "D" in s and "viz" in s and "Kur" not in s and len(s) < 7: rename[c] = "Doviz"
        elif "İD Kuru" in s or "ID Kuru" in s:      rename[c] = "ID Kuru"
        elif "İ.Döviz Tutarı" in s or "I.Doviz Tutari" in s: rename[c] = "ID Tutar"
        elif "İ.Döviz Bakiye" in s or "I.Doviz Bakiye" in s: rename[c] = "ID Bakiye"
        elif i in fallback and c not in rename.values(): rename[c] = fallback[i]
    df = df.rename(columns=rename)

    df = df[df.get("Hesap Kodu", pd.Series()).notna()].copy()
    df = df[df["Hesap Kodu"].astype(str).str.strip() != ""].copy()
    log.info(f"[VNK] Ham satir: {len(df)}")
    return df


def vnk_donustur(df: pd.DataFrame, log) -> pd.DataFrame:
    # Leaf filtre
    kodlar = set(df["Hesap Kodu"].astype(str).str.strip().unique())
    leaf   = leaf_hesap_bul(kodlar)
    ustler = kodlar - leaf
    if ustler:
        log.info(f"[VNK] Filtrelenen ust hesaplar: {sorted(ustler)}")
    df = df[df["Hesap Kodu"].astype(str).str.strip().isin(leaf)].copy()
    log.info(f"[VNK] Leaf sonrasi: {len(df)} satir, {len(leaf)} hesap")

    parsed     = df["Açıklama"].apply(_parse_aciklama_vnk)
    df["_bt"]  = parsed.apply(lambda d: d["belge_turu"])
    df["_ac"]  = parsed.apply(lambda d: d["aciklama_temiz"])
    df["_kh"]  = parsed.apply(lambda d: d["karsi_hesap_kodu"])
    df["_bt"]  = df.apply(lambda r: r["_bt"] if r["_bt"] else str(r.get("Fiş Türü", "") or ""), axis=1)
    df["_tar"] = df["Tarih"].apply(tarih_temizle)
    df["_b"]   = df["Borç"].apply(sayi_temizle)
    df["_a"]   = df["Alacak"].apply(sayi_temizle)
    df["_bk"]  = df["Bakiye"].apply(sayi_temizle)

    # FCY kolonları — varsa kullan, yoksa TRY değerleriyle doldur
    has_fcy = "ID Kuru" in df.columns and "ID Tutar" in df.columns
    if has_fcy:
        df["_kur"]   = df["ID Kuru"].apply(sayi_temizle).replace(0.0, 1.0)
        df["_fcy_t"] = df["ID Tutar"].apply(sayi_temizle)
        df["_fcy_b"] = df["ID Bakiye"].apply(sayi_temizle) if "ID Bakiye" in df.columns else df["_bk"]
        df["_fcy_borc"]   = df.apply(lambda r: abs(r["_fcy_t"]) if r["_b"] > 0 else 0.0, axis=1)
        df["_fcy_alacak"] = df.apply(lambda r: abs(r["_fcy_t"]) if r["_a"] > 0 else 0.0, axis=1)
        log.info(f"[VNK] FCY kolonları bulundu, doviz kuru yazılıyor.")
    else:
        df["_kur"]        = 1.0
        df["_fcy_borc"]   = df["_b"]
        df["_fcy_alacak"] = df["_a"]
        df["_fcy_b"]      = df["_bk"]

    out = _sablon_df(
        tam_kod     = df["Hesap Kodu"].astype(str).str.strip(),
        hesap_adi   = df["Hesap Adı"].astype(str).str.strip(),
        karsi_kod   = df["_kh"],
        tarih       = df["_tar"],
        belge_no    = df["Fiş No"].astype(str).str.strip(),
        belge_turu  = df["_bt"],
        aciklama    = df["_ac"],
        borc        = df["_b"],
        alacak      = df["_a"],
        bakiye      = df["_bk"],
        doviz_kuru  = df["_kur"],
        fcy_borc    = df["_fcy_borc"],
        fcy_alacak  = df["_fcy_alacak"],
        fcy_bakiye  = df["_fcy_b"],
    )
    return out


# ─────────────────────────────────────────────
# WOW FORMAT — HAM VERİ OKUMA
# ─────────────────────────────────────────────
_HESAP_BASLIK_RE  = re.compile(r"^(\d{3}(?:\.\d{2,3})*(?:\.[A-Z0-9]+)*)\s+(.*)", re.IGNORECASE)
_ACIKLAMA_TEMIZ_RE = re.compile(r"^\d{2}/\d{2}/\d{4}-[^,]*,\s*")
_SKIP_WOW = ("Nakli Yekün", "Genel Toplam", "TARİH", "tarih")
_TUTAR_TOL = 0.01


def wow_oku(path: Path, log) -> pd.DataFrame:
    log.info(f"[WOW] Okuma: {path.name}")
    rows, _ = _lese_rows(path)
    log.info(f"[WOW] Ham satir: {len(rows)}")

    kayitlar, guncel_kod, guncel_ad = [], None, None
    for row in rows:
        a = str(row[0]).strip() if row[0] else ""
        m = _HESAP_BASLIK_RE.match(a)
        if m and not any(k in a for k in _SKIP_WOW):
            guncel_kod, guncel_ad = m.group(1).strip(), m.group(2).strip()
            continue
        if not a or any(k in a for k in _SKIP_WOW):
            continue
        tarih, tip, fis_no, aclm = row[0], row[1], row[2], row[3]
        borc, alacak, bakiye      = row[4], row[5], row[6]
        has_tutar = borc not in (None, "0.0", 0.0, 0, "") or alacak not in (None, "0.0", 0.0, 0, "")
        has_tarih = tarih and str(tarih).strip() not in ("", "None")
        if not (has_tarih or has_tutar):
            continue
        kayitlar.append({
            "_hesap_kod": guncel_kod, "_hesap_ad": guncel_ad,
            "Tarih": tarih, "Fiş Türü": tip, "Fiş No": fis_no,
            "Açıklama": aclm, "Borç": borc, "Alacak": alacak, "Bakiye": bakiye,
        })
    df = pd.DataFrame(kayitlar)
    log.info(f"[WOW] Parse sonrasi: {len(df)} satir, {df['_hesap_kod'].nunique()} hesap")
    return df


def wow_donustur(df: pd.DataFrame, log) -> pd.DataFrame:
    df = df[df["_hesap_kod"].notna()].copy()
    kodlar = set(df["_hesap_kod"].astype(str).str.strip().unique())
    leaf   = leaf_hesap_bul(kodlar)
    ustler = kodlar - leaf
    if ustler:
        log.info(f"[WOW] Filtrelenen ust hesaplar: {sorted(ustler)}")
    df = df[df["_hesap_kod"].astype(str).str.strip().isin(leaf)].copy()
    log.info(f"[WOW] Leaf sonrasi: {len(df)} satir")

    df["_tar"] = df["Tarih"].apply(tarih_temizle)
    df["_b"]   = df["Borç"].apply(sayi_temizle)
    df["_a"]   = df["Alacak"].apply(sayi_temizle)
    df["_bk"]  = df["Bakiye"].apply(sayi_temizle)

    def _ac_temiz(v):
        if not v or pd.isna(v): return ""
        return _ACIKLAMA_TEMIZ_RE.sub("", str(v).strip()).strip(", ")

    def _bt(tip, aclm):
        if tip and str(tip).strip() not in ("None", ""):
            return str(tip).strip()
        m2 = re.match(r"^\d{2}/\d{2}/\d{4}-([^,]+)", str(aclm or ""))
        return m2.group(1).strip() if m2 else ""

    df["_ac"] = df["Açıklama"].apply(_ac_temiz)
    df["_bt"] = df.apply(lambda r: _bt(r["Fiş Türü"], r["Açıklama"]), axis=1)

    # Karşı hesap: fiş bazında tutar eşleşmesi
    fis_lookup = {fn: g for fn, g in df.groupby(df["Fiş No"].astype(str).str.strip())}

    def _karsi(fis_no, hesap, borc, alacak):
        grp = fis_lookup.get(str(fis_no).strip())
        if grp is None or len(grp) <= 1: return ""
        diger = grp[grp["_hesap_kod"] != hesap]
        if diger.empty: return ""
        karsilar = set()
        if borc > _TUTAR_TOL:
            es = diger[abs(diger["_a"] - borc) <= _TUTAR_TOL]
            karsilar.update(es["_hesap_kod"].astype(str).str.strip().unique())
        if alacak > _TUTAR_TOL:
            es = diger[abs(diger["_b"] - alacak) <= _TUTAR_TOL]
            karsilar.update(es["_hesap_kod"].astype(str).str.strip().unique())
        if not karsilar:
            karsilar.update(diger["_hesap_kod"].astype(str).str.strip().unique())
        return " | ".join(sorted(karsilar))

    df["_kh"] = df.apply(
        lambda r: _karsi(r["Fiş No"], r["_hesap_kod"], r["_b"], r["_a"]), axis=1
    )
    eslesme = (df["_kh"] != "").sum()
    log.info(f"[WOW] Karsi hesap bulunan: {eslesme}/{len(df)}")

    out = _sablon_df(
        tam_kod    = df["_hesap_kod"].astype(str).str.strip(),
        hesap_adi  = df["_hesap_ad"].astype(str).str.strip(),
        karsi_kod  = df["_kh"],
        tarih      = df["_tar"],
        belge_no   = df["Fiş No"].astype(str).str.strip(),
        belge_turu = df["_bt"],
        aciklama   = df["_ac"],
        borc       = df["_b"],
        alacak     = df["_a"],
        bakiye     = df["_bk"],
    )
    return out


# ─────────────────────────────────────────────
# ORTAK ŞABLON DATAFRAME İNŞAATÇISI
# ─────────────────────────────────────────────
def _sablon_df(**kw) -> pd.DataFrame:
    """
    25 sütunlu standart şablon DataFrame'i oluşturur.
    Sınıf / Alt Hesap kolonları boş bırakılır (hesap kodu parçalanmıyor).
    FCY kolonları opsiyonel — verilmezse TRY değerleriyle doldurulur.
    """
    borc   = kw["borc"]
    alacak = kw["alacak"]
    bakiye = kw["bakiye"]

    out = pd.DataFrame()
    out["Sınıf\nClass"]                                     = ""
    out["Ana Hesap Kodu\nMain Account"]                     = ""
    out["Alt Hesap 1\nSub-Account 1"]                       = ""
    out["Alt Hesap 2\nSub-Account 2"]                       = ""
    out["Alt Hesap 3\nSub-Account 3"]                       = ""
    out["Alt Hesap 4+\nSub-Account 4+"]                     = ""
    out["Tam Hesap Kodu\nFull Account Code"]                = kw["tam_kod"].values
    out["Hesap Adı (TR)\nAccount Name TR"]                  = kw["hesap_adi"].values
    out["Hesap Adı (EN)\nAccount Name EN"]                  = ""
    out["Karşı Hesap Kodu\nCounter Acc. Code"]              = kw["karsi_kod"].values
    out["Karşı Hesap Açıklaması\nCounter Acc. Description"] = ""
    out["Hesap Tipi\nAcc. Type"]                            = kw["tam_kod"].apply(hesap_tipi).values
    out["İşlem Tarihi\nTrans. Date"]                        = kw["tarih"].values
    out["Belge No\nDoc. Number"]                            = kw["belge_no"].values
    out["Belge Türü\nDoc. Type"]                            = kw["belge_turu"].values
    out["Açıklama\nDescription"]                            = kw["aciklama"].values
    out["Borç (DR)\nDebit"]                                 = borc.values
    out["Alacak (CR)\nCredit"]                              = alacak.values
    out["Bakiye\nBalance"]                                  = bakiye.values
    out["Döviz Kuru\nExchange Rate"]                        = kw.get("doviz_kuru", pd.Series([DOVIZ_KURU] * len(borc))).values
    out["Dövizli Borç\nFCY Debit"]                         = kw.get("fcy_borc",   borc).values
    out["Dövizli Alacak\nFCY Credit"]                      = kw.get("fcy_alacak", alacak).values
    out["Dövizli Bakiye\nFCY Balance"]                      = kw.get("fcy_bakiye", bakiye).values
    out["Maliyet Merkezi\nCost Center"]                     = ""
    out["Kaynak Sistem\nSource System"]                     = KAYNAK_SISTEM
    return out


# ─────────────────────────────────────────────
# ÇIKTI YAZMA
# ─────────────────────────────────────────────
def cikti_yaz(df_out: pd.DataFrame, sablon_path: Path, cikti_path: Path, log):
    log.info(f"Cikti yaziliyor: {cikti_path.name}")
    shutil.copy(sablon_path, cikti_path)
    wb = load_workbook(cikti_path)
    sheet_name = next((n for n in wb.sheetnames if "uavin" in n.lower()), None)
    if not sheet_name:
        raise ValueError("Şablonda Muavin Defter sayfası bulunamadı")
    ws = wb[sheet_name]

    DATA_START = 7
    if ws.max_row >= DATA_START:
        ws.delete_rows(DATA_START, ws.max_row - DATA_START + 1)

    thin = Side(style="thin", color="D0D0D0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = list(df_out.columns)

    SAYI_SUTUNLAR = {
        "Borç (DR)\nDebit", "Alacak (CR)\nCredit", "Bakiye\nBalance",
        "Döviz Kuru\nExchange Rate", "Dövizli Borç\nFCY Debit",
        "Dövizli Alacak\nFCY Credit", "Dövizli Bakiye\nFCY Balance",
    }

    for r_idx in range(len(df_out)):
        for c_idx, col in enumerate(cols, start=1):
            val  = df_out.iloc[r_idx, c_idx - 1]
            cell = ws.cell(row=r_idx + DATA_START, column=c_idx)

            if col == "İşlem Tarihi\nTrans. Date":
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    cell.value = val
                    cell.number_format = "DD.MM.YYYY"
                else:
                    cell.value = None
            elif col in SAYI_SUTUNLAR:
                try:
                    cell.value = float(val) if val is not None and str(val) not in ("", "nan") else 0.0
                    cell.number_format = "#,##0.00"
                except Exception:
                    cell.value = val
            else:
                cell.value = val if val is not None and str(val) not in ("None", "nan", "") else ""

            cell.fill = PatternFill("solid", fgColor="F8F9FA" if r_idx % 2 == 0 else "FFFFFF")
            cell.border = brd
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for r in range(DATA_START, DATA_START + len(df_out)):
        ws.row_dimensions[r].height = 18

    col_widths = {1:8,2:12,3:14,4:14,5:14,6:14,7:18,8:38,9:28,10:18,
                  11:28,12:12,13:14,14:18,15:22,16:50,17:14,18:14,
                  19:14,20:12,21:14,22:14,23:14,24:18,25:16}
    for n, w in col_widths.items():
        ws.column_dimensions[get_column_letter(n)].width = w

    wb.save(cikti_path)
    log.info(f"Kaydedildi: {cikti_path}")


# ─────────────────────────────────────────────
# ANA PIPELINE — n8n çağırır
# ─────────────────────────────────────────────
def pipeline(input_path: str, output_path: str, sablon_path: str) -> dict:
    inp     = Path(input_path)
    out     = Path(output_path)
    sablon  = Path(sablon_path)
    log_p   = out.with_suffix(".log")

    log = _setup_logging(log_p)
    log.info(f"Pipeline basladı: {inp.name}")

    if not inp.exists():
        raise FileNotFoundError(f"Girdi dosyası bulunamadı: {inp}")
    if not sablon.exists():
        raise FileNotFoundError(f"Şablon dosyası bulunamadı: {sablon}")

    fmt = format_tespit(inp)
    log.info(f"Tespit edilen format: {fmt}")

    if fmt == "VNK":
        df_ham  = vnk_oku(inp, log)
        df_out  = vnk_donustur(df_ham, log)
        rows_in = len(df_ham)
    elif fmt == "WOW":
        df_ham  = wow_oku(inp, log)
        df_out  = wow_donustur(df_ham, log)
        rows_in = len(df_ham)
    else:
        raise ValueError(f"Tanımlanamayan muavin formatı: {inp.name}")

    cikti_yaz(df_out, sablon, out, log)

    leaf_accounts = sorted(df_out["Tam Hesap Kodu\nFull Account Code"].unique().tolist())

    sonuc = {
        "status":         "success",
        "format":         fmt,
        "input_path":     str(inp),
        "output_path":    str(out),
        "rows_in":        rows_in,
        "rows_out":       len(df_out),
        "leaf_accounts":  leaf_accounts,
        "account_count":  len(leaf_accounts),
        "toplam_borc":    round(float(df_out["Borç (DR)\nDebit"].sum()), 2),
        "toplam_alacak":  round(float(df_out["Alacak (CR)\nCredit"].sum()), 2),
        "log_path":       str(log_p),
        "error":          None,
    }
    log.info(f"Tamamlandı. rows_in={rows_in}, rows_out={len(df_out)}")
    return sonuc


# ─────────────────────────────────────────────
# CLI GİRİŞİ — n8n Execute Command buraya çağırır
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Muavin Pipeline — n8n entegrasyon noktası")
    parser.add_argument("--input",  required=True, help="Ham muavin dosyası (.xls / .xlsx)")
    parser.add_argument("--output", required=True, help="Çıktı dosyası yolu (.xlsx)")
    parser.add_argument("--sablon", required=True, help="Şablon dosyası yolu (.xlsx)")
    args = parser.parse_args()

    try:
        sonuc = pipeline(args.input, args.output, args.sablon)
    except Exception as e:
        sonuc = {
            "status":  "error",
            "format":  None,
            "input_path":  args.input,
            "output_path": args.output,
            "rows_in":  0,
            "rows_out": 0,
            "leaf_accounts":  [],
            "account_count":  0,
            "toplam_borc":    0,
            "toplam_alacak":  0,
            "log_path": "",
            "error":    str(e),
            "traceback": traceback.format_exc(),
        }

    # stdout'a tek satır JSON — n8n bu çıktıyı okur
    print(json.dumps(sonuc, ensure_ascii=False, default=str))
    sys.exit(0 if sonuc["status"] == "success" else 1)


if __name__ == "__main__":
    main()
